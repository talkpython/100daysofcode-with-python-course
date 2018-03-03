#!python3

from openpyxl import load_workbook

wb = load_workbook('Financial Sample.xlsx')
ws1 = wb['Finances 2017']
maxrow = ws1.max_row

def insert_sum():
    ws1['L' + str(maxrow)] = "=SUM(L2:L" + str(maxrow - 1) + ")"
  

if __name__ == "__main__":
    insert_sum()
    wb.save('Financial Sample.xlsx')
